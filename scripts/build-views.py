#!/usr/bin/env python3
"""
build-views.py — regenerate internal/tracker.html and internal/gantt.html
from ticoprojectplanv1.xlsx (the single source of truth).

Usage:
    python build-views.py [plan.xlsx] [out_dir]
      plan.xlsx  default: ticoprojectplanv1.xlsx (same folder)
      out_dir    default: internal/

Reads the 'Project Plan' sheet (and 'Phase Summary' for phase metadata)
and writes two self-contained Brand v1.0 pages (Hebrew, RTL):
    tracker.html  — task list grouped by phase, status pills, notes
    gantt.html    — horizontal timeline, bars colored by status, today marker

Re-run after every xlsx edit, then `git add internal/ && git commit && git push`.

The xlsx is the ONLY source you edit by hand. These two HTML files are
generated artifacts — do not hand-edit them; your changes will be overwritten
on the next build.
"""

import sys
import datetime as dt
from pathlib import Path
import html as _html

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not installed. Run: pip install openpyxl --break-system-packages")

# ── Status code → Brand v1.0 colour (matches book ch6 / site.css pill palette) ──
STATUS = {
    "KEY":      ("#B85A3E", "אבן דרך"),
    "DESIGN":   ("#7C6FCD", "עיצוב"),
    "CONTENT":  ("#B85A3E", "תוכן"),
    "DEV":      ("#4ECDC4", "פיתוח"),
    "IMPL":     ("#4ECDC4", "יישום"),
    "QA":       ("#F8B743", "בדיקות"),
    "FIX":      ("#E05A5F", "תיקונים"),
    "DOCS":     ("#7C6FCD", "תיעוד"),
    "PRINT":    ("#4ECDC4", "הדפסה"),
    "HW":       ("#6B7588", "חומרה"),
    "SUBMIT":   ("#B85A3E", "הגשה"),
    "SETUP":    ("#7C6FCD", "הקמה"),
    "LIVE":     ("#B85A3E", "תערוכה"),
    "TEARDOWN": ("#6B7588", "פירוק"),
}
DEFAULT_COLOR = "#6B7588"

E = lambda s: _html.escape(str(s)) if s is not None else ""


def hex_to_rgba(hexcol, alpha):
    h = hexcol.lstrip("#")
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    return f"rgba({r},{g},{b},{alpha})"


# ── Read the workbook ──────────────────────────────────────────────────────
def read_plan(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Project Plan"]
    phases = []          # list of {name, tasks:[...]}
    current = None
    for r in range(5, ws.max_row + 1):
        a = ws.cell(r, 1).value          # # or phase header
        phase = ws.cell(r, 2).value      # פאזה
        task = ws.cell(r, 3).value       # משימה
        if a is None and task is None:
            continue
        # Phase header row: col A has text, col B (phase) empty
        if phase is None and isinstance(a, str):
            current = {"name": a.strip(), "tasks": []}
            phases.append(current)
            continue
        # Task row
        if current is None:
            current = {"name": "", "tasks": []}
            phases.append(current)
        current["tasks"].append({
            "num":    a,
            "phase":  phase,
            "task":   task,
            "start":  ws.cell(r, 4).value,
            "end":    ws.cell(r, 5).value,
            "days":   ws.cell(r, 6).value,
            "status": (ws.cell(r, 7).value or "").strip(),
            "notes":  ws.cell(r, 8).value,
        })
    # title/meta (normalise old t.Co. → Odysee)
    title = (ws.cell(1, 1).value or "Odysee — Project Plan")
    title = title.replace("t.Co.", "Odysee").replace("t.Co", "Odysee")
    meta = (ws.cell(2, 1).value or "")
    return phases, title, meta


def all_tasks(phases):
    return [t for p in phases for t in p["tasks"]]


# ── Shared HTML head + header ───────────────────────────────────────────────
def head(title, active):
    nav = [("../index.html", "Home", False),
           ("tracker.html", "Tracker", active == "tracker"),
           ("gantt.html", "Gantt", active == "gantt")]
    navhtml = "\n".join(
        f'      <li><a href="{u}"{" class=\"active\"" if act else ""}>{n}</a></li>'
        for u, n, act in nav)
    roundel = (
        '<svg class="bm__o" viewBox="0 0 100 100" aria-hidden="true">'
        '<circle class="bm-inner" cx="50" cy="50" r="45.4"/>'
        '<circle class="bm-ring" cx="50" cy="50" r="45.4" fill="none" stroke-width="9.2"/>'
        '<g transform="rotate(-32 50 50)">'
        '<path class="bm-trail" d="M 26 50 A 24 24 0 0 0 74 50" fill="none" stroke-width="6.5" '
        'stroke-linecap="round" stroke-dasharray="0.1 12"/>'
        '<circle class="bm-start" cx="26" cy="50" r="3.6"/>'
        '<circle class="bm-end" cx="74" cy="50" r="5.4"/>'
        '</g></svg>')
    return f"""<!DOCTYPE html>
<html lang="he" dir="rtl" data-theme="system">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{E(title)}</title>
<meta name="robots" content="noindex">
<link rel="stylesheet" href="../brand/odysee-tokens.css">
<link rel="stylesheet" href="../assets/css/site.css">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=Sora:wght@400;500;700;800&family=Heebo:wght@400;500;700;800&family=JetBrains+Mono:wght@400;500&display=swap">
<style>
{COMPONENT_CSS}
</style>
</head>
<body>
<div class="theme-bar">
  <button class="theme-btn" onclick="setTheme('light')" id="btn-light">☀ בהיר</button>
  <button class="theme-btn" onclick="setTheme('system')" id="btn-system">⊙ מערכת</button>
  <button class="theme-btn" onclick="setTheme('dark')" id="btn-dark">● כהה</button>
</div>
<header class="site-header" dir="ltr">
  <div class="iv-brand">
    <a href="../index.html" class="bm" aria-label="Odysee — home">
      {roundel}<span class="bm__text">dysee</span>
    </a>
    <span class="iv-tagline">Internal · {E(active.title())}</span>
  </div>
  <nav><ul class="site-nav">
{navhtml}
  </ul></nav>
</header>
"""


FOOT = """<footer class="site-footer">
  <div class="site-footer__left">David Castiel · M.Design · HIT 2026 — Internal view, generated from xlsx</div>
  <div class="site-footer__right" dir="ltr">Generated %s</div>
</footer>
<script>
const root=document.documentElement,mq=matchMedia('(prefers-color-scheme: dark)');
let cur=localStorage.getItem('odysee-theme')||'system';
function apply(t){root.setAttribute('data-theme',t==='system'?(mq.matches?'dark':'light'):t);
  ['light','system','dark'].forEach(x=>document.getElementById('btn-'+x).classList.toggle('active',x===t));
  cur=t;localStorage.setItem('odysee-theme',t);}
function setTheme(t){apply(t);}
mq.addEventListener('change',()=>{if(cur==='system')apply('system');});
apply(cur);
</script>
</body>
</html>"""


COMPONENT_CSS = """
.iv-brand{display:flex;flex-direction:column;gap:4px}
.bm{display:inline-flex;align-items:center;gap:2px;text-decoration:none;line-height:.9}
.bm__o{width:2.2em;height:2.2em}
.bm__text{font-family:var(--od-font-sans,'Sora',sans-serif);font-weight:800;font-size:34px;
  letter-spacing:-.035em;color:var(--od-fg)}
.bm-inner{fill:#fff}.bm-ring{stroke:var(--od-navy)}.bm-trail{stroke:var(--od-navy)}
.bm-start{fill:var(--od-navy)}.bm-end{fill:var(--od-red)}
[data-theme="dark"] .bm-ring{stroke:var(--od-paper)}
.iv-tagline{font-size:10px;letter-spacing:.22em;text-transform:uppercase;
  color:var(--od-fg-subtle);font-weight:500;margin-top:2px}
.iv-wrap{max-width:1200px;margin:0 auto;padding:32px 24px 64px}
.iv-h1{font-family:var(--od-font-sans,'Sora',sans-serif);font-size:28px;font-weight:800;
  letter-spacing:-.02em;color:var(--od-fg);margin-bottom:4px}
.iv-meta{font-size:12px;color:var(--od-fg-subtle);margin-bottom:8px;direction:ltr;text-align:left}
.iv-summary{display:flex;gap:10px;flex-wrap:wrap;margin:18px 0 28px}
.iv-stat{background:var(--od-bg-sunken);border:1px solid var(--od-border);
  border-radius:var(--od-r-md,10px);padding:12px 18px;min-width:96px}
.iv-stat .n{font-family:var(--od-font-sans,'Sora',sans-serif);font-size:24px;font-weight:800;color:var(--od-fg)}
.iv-stat .l{font-size:11px;color:var(--od-fg-subtle);letter-spacing:.04em}
.iv-phase{margin-top:30px}
.iv-phase h2{font-family:var(--od-font-sans,'Sora',sans-serif);font-size:16px;font-weight:700;
  color:var(--od-fg);margin-bottom:2px}
.iv-phase .sub{font-size:11px;color:var(--od-fg-subtle);direction:ltr;text-align:right;margin-bottom:12px}
/* pills */
.pill{display:inline-block;padding:2px 10px;border-radius:999px;font-size:9px;font-weight:600;
  letter-spacing:.10em;text-transform:uppercase;font-family:var(--od-font-sans,'Sora',sans-serif);
  white-space:nowrap;direction:ltr}
/* tracker table */
.iv-table{width:100%;border-collapse:collapse;font-size:13px;
  font-family:var(--od-font-sans,'Sora',sans-serif)}
.iv-table th{background:var(--od-navy);color:#fff;font-size:9px;font-weight:600;letter-spacing:.10em;
  text-transform:uppercase;padding:9px 12px;text-align:right}
.iv-table td{padding:9px 12px;color:var(--od-fg-muted);border-bottom:1px solid var(--od-border);
  vertical-align:top}
.iv-table tr:nth-child(even) td{background:var(--od-bg-sunken)}
.iv-table .c-num{font-family:var(--od-font-mono,monospace);font-size:11px;color:var(--od-fg-subtle);direction:ltr}
.iv-table .c-task{color:var(--od-fg);font-weight:500}
.iv-table .c-date{font-family:var(--od-font-mono,monospace);font-size:11px;color:var(--od-fg-subtle);direction:ltr;white-space:nowrap}
.iv-table .c-notes{font-size:12px;color:var(--od-fg-subtle);line-height:1.5}
/* gantt */
.gantt{margin-top:8px;font-family:var(--od-font-sans,'Sora',sans-serif);direction:ltr}
.gantt-months{display:flex;border-bottom:1px solid var(--od-border-strong);
  margin-left:230px;position:relative;height:22px}
.gantt-month{font-size:10px;color:var(--od-fg-subtle);letter-spacing:.06em;
  border-right:1px solid var(--od-border);padding:2px 0 0 4px;box-sizing:border-box;text-align:left}
.gantt-row{display:flex;align-items:center;height:30px;border-bottom:1px solid var(--od-border)}
.gantt-label{width:230px;flex-shrink:0;font-size:12px;color:var(--od-fg);padding:0 10px;
  white-space:nowrap;overflow:hidden;text-overflow:ellipsis;direction:rtl;text-align:right;
  display:flex;align-items:center;gap:8px}
.gantt-label .gtext{overflow:hidden;text-overflow:ellipsis;flex:1}
.gantt-label .gnum{font-family:var(--od-font-mono,monospace);font-size:10px;color:var(--od-fg-subtle);
  direction:ltr;flex-shrink:0;min-width:22px;text-align:left;order:2}
.gantt-track{position:relative;flex:1;height:100%}
.gantt-bar{position:absolute;top:7px;height:16px;border-radius:4px;opacity:.9;
  display:flex;align-items:center;direction:ltr}
.gantt-phase{font-family:var(--od-font-sans,'Sora',sans-serif);font-size:11px;font-weight:700;
  color:var(--od-red);letter-spacing:.06em;text-transform:uppercase;
  padding:14px 10px 4px;border-bottom:1px solid var(--od-border-strong);margin-top:6px;direction:rtl;text-align:right}
.gantt-today{position:absolute;top:0;bottom:0;width:2px;background:var(--od-red);z-index:5}
.gantt-legend{display:flex;gap:14px;flex-wrap:wrap;margin-top:22px;padding-top:16px;
  border-top:1px solid var(--od-border);direction:rtl}
.gantt-legend .lg{display:flex;align-items:center;gap:6px;font-size:11px;color:var(--od-fg-muted)}
.gantt-legend .sw{width:12px;height:12px;border-radius:3px}
@media(max-width:640px){
  .gantt-months{margin-left:130px}.gantt-label{width:130px;font-size:11px}
  .iv-table .c-notes{display:none}
}
"""


def pill(status):
    color, _ = STATUS.get(status, (DEFAULT_COLOR, status))
    bg = hex_to_rgba(color, 0.16)
    return f'<span class="pill" style="background:{bg};color:{color}">{E(status)}</span>'


def fmt(d):
    return d.strftime("%Y-%m-%d") if isinstance(d, dt.datetime) else E(d)


# ── tracker.html ────────────────────────────────────────────────────────────
def build_tracker(phases, title, meta):
    tasks = all_tasks(phases)
    # status counts
    from collections import Counter
    counts = Counter(t["status"] for t in tasks)
    stats = "".join(
        f'<div class="iv-stat"><div class="n">{len(tasks)}</div><div class="l">משימות</div></div>'
        if i == 0 else "" for i in range(1))
    stats += f'<div class="iv-stat"><div class="n">{len(phases)}</div><div class="l">פאזות</div></div>'
    for st, c in counts.most_common(4):
        col = STATUS.get(st, (DEFAULT_COLOR,))[0]
        stats += (f'<div class="iv-stat"><div class="n" style="color:{col}">{c}</div>'
                  f'<div class="l">{E(st)}</div></div>')

    body = [f'<div class="iv-wrap">',
            f'<h1 class="iv-h1">מעקב פרויקט</h1>',
            f'<div class="iv-meta">{E(meta)}</div>',
            f'<div class="iv-summary">{stats}</div>']
    for p in phases:
        if not p["tasks"]:
            continue
        body.append(f'<div class="iv-phase"><h2>{E(p["name"])}</h2>')
        body.append('<table class="iv-table"><thead><tr>'
                     '<th style="width:36px">#</th><th>משימה</th>'
                     '<th style="width:90px">התחלה</th><th style="width:90px">סיום</th>'
                     '<th style="width:50px">ימים</th><th style="width:90px">סטטוס</th>'
                     '<th>הערות</th></tr></thead><tbody>')
        for t in p["tasks"]:
            body.append(
                f'<tr><td class="c-num">{E(t["num"])}</td>'
                f'<td class="c-task">{E(t["task"])}</td>'
                f'<td class="c-date">{fmt(t["start"])}</td>'
                f'<td class="c-date">{fmt(t["end"])}</td>'
                f'<td class="c-num">{E(t["days"])}</td>'
                f'<td>{pill(t["status"])}</td>'
                f'<td class="c-notes">{E(t["notes"])}</td></tr>')
        body.append('</tbody></table></div>')
    body.append('</div>')
    return head("Odysee — מעקב פרויקט", "tracker") + "\n".join(body) + \
        (FOOT % dt.date.today().isoformat())


# ── gantt.html ──────────────────────────────────────────────────────────────
def build_gantt(phases, title, meta):
    tasks = all_tasks(phases)
    starts = [t["start"] for t in tasks if isinstance(t["start"], dt.datetime)]
    ends = [t["end"] for t in tasks if isinstance(t["end"], dt.datetime)]
    p0, p1 = min(starts), max(ends)
    total = (p1 - p0).days or 1

    def pct(d):
        return (d - p0).days / total * 100

    # month headers
    months = []
    cur = dt.datetime(p0.year, p0.month, 1)
    while cur <= p1:
        nxt = dt.datetime(cur.year + (cur.month == 12), (cur.month % 12) + 1, 1)
        seg_start = max(cur, p0)
        seg_end = min(nxt, p1)
        w = (seg_end - seg_start).days / total * 100
        months.append((cur.strftime("%b %Y"), w))
        cur = nxt
    months_html = "".join(
        f'<div class="gantt-month" style="width:{w:.3f}%">{m}</div>' for m, w in months)

    today = dt.datetime.now()
    today_marker = ""
    if p0 <= today <= p1:
        today_marker = (f'<div class="gantt-today" style="left:{pct(today):.3f}%" '
                        f'title="today {today.date()}"></div>')

    rows = []
    for p in phases:
        if not p["tasks"]:
            continue
        rows.append(f'<div class="gantt-phase">{E(p["name"])}</div>')
        for t in p["tasks"]:
            if not (isinstance(t["start"], dt.datetime) and isinstance(t["end"], dt.datetime)):
                continue
            color = STATUS.get(t["status"], (DEFAULT_COLOR,))[0]
            left = pct(t["start"])
            width = max((t["end"] - t["start"]).days / total * 100, 0.8)
            label = E(t["task"])
            rows.append(
                f'<div class="gantt-row"><div class="gantt-label">'
                f'<span class="gnum">{E(t["num"])}</span><span class="gtext">{label}</span></div>'
                f'<div class="gantt-track">'
                f'<div class="gantt-bar" style="left:{left:.3f}%;width:{width:.3f}%;'
                f'background:{color}" title="{label} · {fmt(t["start"])}→{fmt(t["end"])}"></div>'
                f'</div></div>')

    # legend (unique statuses present)
    seen = []
    for t in tasks:
        if t["status"] not in seen:
            seen.append(t["status"])
    legend = "".join(
        f'<div class="lg"><span class="sw" style="background:{STATUS.get(s,(DEFAULT_COLOR,))[0]}"></span>{E(s)}</div>'
        for s in seen)

    body = [
        '<div class="iv-wrap">',
        '<h1 class="iv-h1">Gantt — ציר זמן</h1>',
        f'<div class="iv-meta">{E(meta)}</div>',
        f'<div class="iv-meta" dir="ltr">{p0.date()} → {p1.date()} · {total} days</div>',
        '<div class="gantt">',
        f'<div class="gantt-months">{months_html}{today_marker}</div>',
        "".join(rows),
        f'<div class="gantt-legend">{legend}</div>',
        '</div></div>',
    ]
    return head("Odysee — Gantt", "gantt") + "\n".join(body) + \
        (FOOT % dt.date.today().isoformat())


# ── main ────────────────────────────────────────────────────────────────────
def main():
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("ticoprojectplanv1.xlsx")
    out = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("internal")
    if not xlsx.exists():
        sys.exit(f"xlsx not found: {xlsx}")
    out.mkdir(parents=True, exist_ok=True)

    phases, title, meta = read_plan(xlsx)
    n = sum(len(p["tasks"]) for p in phases)

    (out / "tracker.html").write_text(build_tracker(phases, title, meta), encoding="utf-8")
    (out / "gantt.html").write_text(build_gantt(phases, title, meta), encoding="utf-8")

    print(f"✓ Read {n} tasks across {len([p for p in phases if p['tasks']])} phases from {xlsx.name}")
    print(f"✓ Wrote {out/'tracker.html'}")
    print(f"✓ Wrote {out/'gantt.html'}")
    print("\nNext: git add internal/ && git commit -m 'chore: rebuild project views' && git push")


if __name__ == "__main__":
    main()
